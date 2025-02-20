import time
import datetime
from SSVEP_Feedback.cfg import *
import openpyxl
from openpyxl.styles import Alignment
import random
import os
import json
from typing import List


class DialogueManager:
    def __init__(self):
        self.text_his = ['']
        self.formatted_dialogue = []

    def setText(self, text):
        if text == self.text_his[-1] == '':
            return
        self.text_his.append(text)

    def getText(self):
        return self.text_his[-1]

    def setFormattedUtterance(self, formatted_utterance: List):
        self.formatted_dialogue.append([formatted_utterance[0].strip(),formatted_utterance[1].strip()])

    def getUtterance(self):
        return ": ".join(self.formatted_dialogue[-1])

    def getUtteranceContent(self):
        return self.formatted_dialogue[-1][1]

    def setDialogue(self, dialogue: str):
        dialogue = dialogue.strip()
        if dialogue == '':
            return
        for utterance in dialogue.split('\n'):
            if utterance:
                idx = utterance.find(':')
                self.formatted_dialogue.append([utterance[:idx], utterance[idx + 1:].strip()])

    def setFormattedDialogue(self, formattedDialogue):
        self.formatted_dialogue.extend(formattedDialogue)

    def getDialogue(self):
        return "\n".join(": ".join(formatted_utterance) for formatted_utterance in self.formatted_dialogue)

    def getFormattedDialogue(self):
        return self.formatted_dialogue

    def undo(self):
        if len(self.text_his) > 1:
            text = self.text_his.pop(-1)
        else:
            text = self.text_his[0]
        return text

    def clear(self):
        self.text_his = ['']
        self.formatted_dialogue = []


class DialogueExp:
    def __init__(self):
        self.testSamples = []
        self.idx = -1
        self.datetime = None

    def createTestSamples(self):
        workbook = openpyxl.load_workbook(DATASET_FILE)
        sheet = workbook.active
        assert sheet.cell(row=1, column=1).value == 'context'
        assert sheet.cell(row=1, column=2).value == 'target utterance'
        for r in range(2, sheet.max_row + 1):
            if (sheet.cell(row=r, column=2).value) is None:
                continue
            context = sheet.cell(row=r, column=1).value
            target_message = sheet.cell(row=r, column=2).value
            self.testSamples.append({'context':context if context else '', 'target utterance': target_message, 'runTime': [], 'strokes': []})
        random.shuffle(self.testSamples)
        self.testSamples = self.testSamples[:RUN_NUM]

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(1, 1).value = 'context'
        sheet.cell(1, 2).value = 'target utterance'
        for r in range(2, len(self.testSamples) + 2):
            sheet.cell(r, 1).value = self.testSamples[r - 2]['context']
            sheet.cell(r, 1).alignment = Alignment(wrapText=True)
            sheet.cell(r, 2).value = self.testSamples[r - 2]['target utterance']
            sheet.cell(r, 2).alignment = Alignment(wrapText=True)
        workbook.save('dialogue samples.xlsx')
        self.datetime = str(datetime.datetime.now()).split('.')[0].replace(' ', '_').replace(':', '-')

    def getOneSample(self):
        self.idx += 1
        return self.idx, self.testSamples[self.idx]

    def setSampleUtterance(self, utterance, idx = None):
        self.testSamples[idx if idx else self.idx]['user utterance'] = utterance

    def startRunTime(self, idx = None):
        self.testSamples[idx if idx else self.idx]['runTime'].append(time.time())

    def stopRunTime(self, idx = None):
        self.testSamples[idx if idx else self.idx]['runTime'].append(time.time())

    def addStoke(self, stroke_idx, idx = None):
        self.testSamples[idx if idx else self.idx]['strokes'].append(stroke_idx)

    def saveTestResults(self):
        result_dir = f'SSVEP_Feedback/output/{SUBJECT_NAME}'
        os.makedirs(result_dir, exist_ok=True)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(1, 1).value = 'context'
        sheet.cell(1, 2).value = 'target utterance'
        sheet.cell(1, 3).value = 'user utterance'
        sheet.cell(1, 4).value = 'time length'
        sheet.cell(1, 5).value = 'strokes'
        for r in range(2, len(self.testSamples) + 2):
            sheet.cell(r, 1).value = self.testSamples[r - 2]['context']
            sheet.cell(r, 1).alignment = Alignment(wrapText=True)
            sheet.cell(r, 2).value = self.testSamples[r - 2]['target utterance']
            sheet.cell(r, 2).alignment = Alignment(wrapText=True)
            if 'user utterance' in self.testSamples[r - 2]:
                sheet.cell(r, 3).value = self.testSamples[r - 2]['user utterance']
                sheet.cell(r, 3).alignment = Alignment(wrapText=True)
                s, e = self.testSamples[r - 2]['runTime']
                sheet.cell(r, 4).value = int(e - s)
                sheet.cell(r, 4).alignment = Alignment(wrapText=True)
                sheet.cell(r, 5).value = json.dumps(self.testSamples[r - 2]['strokes'])
                sheet.cell(r, 5).alignment = Alignment(wrapText=True)
        file = DATASET_FILE.split('\\')[-1][:-5] + ' results.xlsx'
        workbook.save(f"{result_dir}/{self.datetime} {file}")